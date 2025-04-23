import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import json
import os
import openpyxl
import sys
import threading
from driver_utils import terminate_chrome_processes, initialize_driver
from excel_utils import read_excel_data
from form_utils import get_form_headers, fill_google_form
from matching_utils import match_headers

# Dynamic resource path for PyInstaller
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller."""
    if hasattr(sys, '_MEIPASS'):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path(__file__).parent
    return str(base_path / relative_path)

# Application data directory
APP_DATA_DIR = Path(os.getenv("APPDATA")) / "TRC_AUTO"
APP_DATA_DIR.mkdir(exist_ok=True)
CONFIG_JSON = str(APP_DATA_DIR / "config.json")

class ConfigGUI:
    """GUI for managing configuration settings stored in config.json."""
    def __init__(self, root):
        self.root = root
        self.root.title("Google Form Automation - FBB Dept")
        self.root.resizable(False, False)
        self.entries = {}
        self.is_running = False
        self.workbook = None  # Store workbook for access during cleanup
        self.load_config()
        
        # Center the window on the screen
        window_width = 700
        window_height = 500
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x_position = (screen_width - window_width) // 2
        y_position = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
        
        self.create_widgets()
        self.apply_styles()

    def apply_styles(self):
        """Apply custom styles for a good and cute GUI."""
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", font=("Helvetica", 10), padding=5)
        style.configure("TButton", font=("Helvetica", 10, "bold"), padding=5, background="#4CAF50")
        style.configure("TEntry", font=("Helvetica", 10), padding=5)
        style.configure("TFrame", background="#f0f0f0")
        self.root.configure(bg="#f0f0f0")

    def load_config(self):
        """Load configuration from config.json."""
        default_user_data = Path(os.getenv("LOCALAPPDATA")) / "Google" / "Chrome" / "User Data"
        default_config = {
            "GOOGLE_FORM_URL": "",
            "EXCEL_FILE": "",
            "USER_DATA_DIR": str(default_user_data),
            "PROFILE_DIR": "Default",
            "SIMILARITY_THRESHOLD": 80
        }

        try:
            if not Path(CONFIG_JSON).exists():
                with open(CONFIG_JSON, 'w', encoding='utf-8') as f:
                    json.dump(default_config, f, indent=4)
            with open(CONFIG_JSON, 'r', encoding='utf-8') as f:
                self.config_values = json.load(f)
            logging.info(f"Loaded config from {CONFIG_JSON}: {self.config_values}")
        except Exception as e:
            logging.error(f"Failed to load {CONFIG_JSON}: {e}")
            self.config_values = default_config

        for key, value in default_config.items():
            if key not in self.config_values:
                self.config_values[key] = value

        self.account_name, self.email = self.get_account_info()

    def get_account_info(self):
        """Retrieve account name and email from Chrome's Local State file."""
        try:
            local_state_path = Path(self.config_values["USER_DATA_DIR"]) / "Local State"
            if not local_state_path.is_file():
                logging.warning(f"Local State file not found at {local_state_path}")
                return "Not available", "Not available"

            with open(local_state_path, 'r', encoding='utf-8') as f:
                local_state = json.load(f)

            profile_cache = local_state.get("profile", {}).get("info_cache", {})
            profile_info = profile_cache.get(self.config_values["PROFILE_DIR"], {})

            account_name = profile_info.get("name", "Not available")
            email = profile_info.get("user_name", "Not available")

            logging.info(f"Retrieved account info - Name: {account_name}, Email: {email}")
            return account_name, email
        except Exception as e:
            logging.error(f"Failed to read Local State: {e}")
            return "Not available", "Not available"

    def save_config(self):
        """Save configuration to config.json."""
        try:
            config_data = {
                "GOOGLE_FORM_URL": self.entries["GOOGLE_FORM_URL"].get().strip(),
                "EXCEL_FILE": str(Path(self.entries["EXCEL_FILE"].get())) if self.entries["EXCEL_FILE"].get() else "",
                "USER_DATA_DIR": str(Path(self.entries["USER_DATA_DIR"].get())),
                "PROFILE_DIR": self.entries["PROFILE_DIR"].get().strip(),
                "SIMILARITY_THRESHOLD": self.config_values.get("SIMILARITY_THRESHOLD", 80)
            }
            with open(CONFIG_JSON, "w", encoding='utf-8') as f:
                json.dump(config_data, f, indent=4)
            logging.info(f"Configuration saved to {CONFIG_JSON}")
        except Exception as e:
            logging.error(f"Failed to save {CONFIG_JSON}: {e}")
            messagebox.showerror("Error", f"Failed to save configuration: {e}")

    def clear_config(self):
        """Reset config.json to default values."""
        if messagebox.askyesno("Confirm", "Are you sure you want to reset settings to defaults?"):
            try:
                default_user_data = Path(os.getenv("LOCALAPPDATA")) / "Google" / "Chrome" / "User Data"
                default_config = {
                    "GOOGLE_FORM_URL": "",
                    "EXCEL_FILE": "",
                    "USER_DATA_DIR": str(default_user_data),
                    "PROFILE_DIR": "Default",
                    "SIMILARITY_THRESHOLD": 80
                }
                with open(CONFIG_JSON, "w", encoding='utf-8') as f:
                    json.dump(default_config, f, indent=4)
                logging.info(f"Configuration reset to defaults in {CONFIG_JSON}")

                self.config_values = default_config
                for key, entry in self.entries.items():
                    entry.delete(0, tk.END)
                    entry.insert(0, self.config_values[key])

                self.account_name, self.email = "Not available", "Not available"
                self.account_name_label.config(text=f"Account Name: {self.account_name}")
                self.email_label.config(text=f"Email: {self.email}")

                messagebox.showinfo("Success", "Settings reset to defaults")
            except Exception as e:
                logging.error(f"Failed to reset {CONFIG_JSON}: {e}")
                messagebox.showerror("Error", f"Failed to reset configuration: {e}")

    def create_widgets(self):
        """Create GUI widgets for configuration inputs and account info, centered."""
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.columnconfigure(0, weight=1)  # Center content horizontally

        # Logo text (centered)
        logo_label = tk.Label(
            main_frame,
            text="FBB Dept",
            font=("Helvetica", 24, "bold"),
            fg="#4CAF50",
            bg="#f0f0f0"
        )
        logo_label.grid(row=0, column=0, pady=(0, 20), sticky=tk.EW)

        # Subtitle (centered)
        subtitle_label = tk.Label(
            main_frame,
            text="Google Form Automation",
            font=("Helvetica", 12, "italic"),
            fg="#555555",
            bg="#f0f0f0"
        )
        subtitle_label.grid(row=1, column=0, pady=(0, 20), sticky=tk.EW)

        config_fields = [
            ("Google Form URL", "GOOGLE_FORM_URL", False),
            ("Excel File", "EXCEL_FILE", True),
            ("User Data Directory", "USER_DATA_DIR", True),
            ("Profile Directory", "PROFILE_DIR", False),
        ]

        for idx, (label_text, config_key, is_file) in enumerate(config_fields, start=2):
            # Frame for each input row to center content
            row_frame = ttk.Frame(main_frame)
            row_frame.grid(row=idx, column=0, sticky=tk.EW, pady=8)
            row_frame.columnconfigure(1, weight=1)  # Allow entry to expand

            ttk.Label(row_frame, text=label_text + ":").grid(row=0, column=0, sticky=tk.W)
            entry = ttk.Entry(row_frame, width=50)
            entry.insert(0, self.config_values[config_key])
            entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
            self.entries[config_key] = entry
            if is_file:
                browse_btn = ttk.Button(row_frame, text="Browse", command=lambda k=config_key: self.browse_file(k))
                browse_btn.grid(row=0, column=2, padx=5)

        row_offset = len(config_fields) + 2
        # Account info (centered)
        self.account_name_label = ttk.Label(main_frame, text=f"Account Name: {self.account_name}")
        self.account_name_label.grid(row=row_offset, column=0, sticky=tk.EW, pady=8)

        self.email_label = ttk.Label(main_frame, text=f"Email: {self.email}")
        self.email_label.grid(row=row_offset + 1, column=0, sticky=tk.EW, pady=8)

        # Button frame (centered)
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=row_offset + 2, column=0, pady=20, sticky=tk.EW)
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)

        self.save_run_btn = ttk.Button(button_frame, text="Save and Run", command=self.save_and_run)
        self.save_run_btn.grid(row=0, column=0, padx=10)

        ttk.Button(button_frame, text="Clear Settings", command=self.clear_config).grid(row=0, column=1, padx=10)

        # Status label (centered)
        self.status_label = ttk.Label(main_frame, text="Ready", foreground="black")
        self.status_label.grid(row=row_offset + 3, column=0, pady=10, sticky=tk.EW)

    def browse_file(self, config_key):
        """Open file/directory dialog for specific configuration fields and update account info."""
        if config_key == "EXCEL_FILE":
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
            if file_path and Path(file_path).is_file():
                self.entries[config_key].delete(0, tk.END)
                self.entries[config_key].insert(0, str(Path(file_path)))
                self.config_values[config_key] = str(Path(file_path))
                logging.info(f"Selected {config_key}: {file_path}")
            else:
                messagebox.showerror("Error", "Invalid Excel file selected")
        elif config_key == "USER_DATA_DIR":
            dir_path = filedialog.askdirectory()
            if dir_path and Path(dir_path).is_dir():
                self.entries[config_key].delete(0, tk.END)
                self.entries[config_key].insert(0, str(Path(dir_path)))
                self.config_values[config_key] = str(Path(dir_path))
                self.account_name, self.email = self.get_account_info()
                self.account_name_label.config(text=f"Account Name: {self.account_name}")
                self.email_label.config(text=f"Email: {self.email}")
                logging.info(f"Selected {config_key}: {dir_path}")
            else:
                messagebox.showerror("Error", "Invalid directory selected")

    def prevent_close(self):
        """Allow closing the application but warn user to wait for Excel update."""
        if not self.is_running:
            self.root.destroy()
            return

        response = messagebox.askyesno(
            "Confirm Close",
            "Automation is still running. Please wait for the Excel file to be updated. Closing now may result in incomplete data. Do you want to close anyway?"
        )
        if response:
            # Save the workbook if it exists
            if self.workbook is not None:
                try:
                    self.workbook.save(self.config_values["EXCEL_FILE"])
                    logging.info("Excel file saved before closing application")
                except Exception as e:
                    logging.error(f"Failed to save Excel file before closing: {e}")
                    messagebox.showerror("Error", f"Failed to save Excel file: {e}")
            self.root.destroy()

    def save_and_run(self):
        """Validate inputs, save to config.json, and run automation in a separate thread."""
        if self.is_running:
            messagebox.showwarning("Warning", "Automation is already running. Please wait for it to complete.")
            return

        try:
            google_form_url = self.entries["GOOGLE_FORM_URL"].get().strip()
            excel_file = self.entries["EXCEL_FILE"].get().strip()
            user_data_dir = self.entries["USER_DATA_DIR"].get().strip()
            profile_dir = self.entries["PROFILE_DIR"].get().strip()

            logging.info(f"Inputs - Google Form URL: {google_form_url}, Excel File: {excel_file}, User Data Dir: {user_data_dir}, Profile Dir: {profile_dir}")

            if not google_form_url:
                raise ValueError("Google Form URL cannot be empty")
            if not google_form_url.startswith("http"):
                raise ValueError("Invalid Google Form URL: Must start with 'http'")
            if not excel_file:
                raise ValueError("Excel file path cannot be empty")
            excel_path = Path(excel_file)
            if not excel_path.is_file():
                raise ValueError(f"Excel file does not exist: {excel_file}")
            if not user_data_dir:
                raise ValueError("User data directory cannot be empty")
            if not Path(user_data_dir).is_dir():
                raise ValueError(f"User data directory does not exist: {user_data_dir}")
            if not profile_dir:
                raise ValueError("Profile directory cannot be empty")

            self.config_values["GOOGLE_FORM_URL"] = google_form_url
            self.config_values["EXCEL_FILE"] = str(excel_path)
            self.config_values["USER_DATA_DIR"] = user_data_dir
            self.config_values["PROFILE_DIR"] = profile_dir
            self.save_config()

            self.account_name, self.email = self.get_account_info()
            self.account_name_label.config(text=f"Account Name: {self.account_name}")
            self.email_label.config(text=f"Email: {self.email}")

            self.is_running = True
            self.save_run_btn.config(state="disabled")
            self.status_label.config(text="Running, please wait for Excel update...", foreground="red")
            self.root.protocol("WM_DELETE_WINDOW", self.prevent_close)

            threading.Thread(target=self.run_automation, daemon=True).start()

        except ValueError as e:
            logging.error(f"Validation error: {e}")
            messagebox.showerror("Error", str(e))
        except Exception as e:
            logging.error(f"Unexpected error in save_and_run: {e}")
            messagebox.showerror("Error", f"Unexpected error: {e}")

    def run_automation(self):
        """Run the main automation process and display the result."""
        try:
            result = main(self.config_values, self)
            self.root.after(0, lambda: self.show_result(result))
        except Exception as e:
            self.root.after(0, lambda: self.show_result(f"Unexpected error: {e}"))
        finally:
            self.root.after(0, self.reset_gui)

    def show_result(self, result):
        """Display the result of the automation process."""
        google_form_name = self.config_values["GOOGLE_FORM_URL"].split('/')[-2] if '/' in self.config_values["GOOGLE_FORM_URL"] else "Google Form"

        if result == "Success":
            message = f"Automation completed successfully! Excel file has been updated."
            messagebox.showinfo("Success", message)
            self.status_label.config(text="Completed: Excel file updated", foreground="green")
        else:
            message = f"{result}"
            messagebox.showerror("Error", message)
            self.status_label.config(text="Error: Check logs for details", foreground="red")

    def reset_gui(self):
        """Re-enable the GUI after automation completes."""
        self.is_running = False
        self.workbook = None  # Clear workbook reference
        self.save_run_btn.config(state="normal")
        self.root.protocol("WM_DELETE_WINDOW", self.root.destroy)
        if self.status_label.cget("text").startswith("Running"):
            self.status_label.config(text="Ready", foreground="black")

def main(config, gui):
    """Main function to orchestrate the automation process."""
    driver = None
    wb = None
    filepath = Path(config["EXCEL_FILE"])
    try:
        if not filepath:
            raise ValueError("Excel file path is empty")
        if not filepath.is_file():
            raise FileNotFoundError(f"Excel file not found: {filepath}")

        try:
            wb = openpyxl.load_workbook(filepath)
            gui.workbook = wb  # Store workbook in GUI instance for access during cleanup
        except Exception as e:
            logging.error(f"Failed to load Excel file: {e}")
            raise ValueError(f"Failed to load Excel file: {e}")

        sheet = wb.active
        excel_headers, rows = read_excel_data(config["EXCEL_FILE"])
        logging.info(f"Total rows to process: {len(rows)}")

        note_column = None
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value and isinstance(cell.value, str) and cell.value.lower() == "note":
                note_column = col_idx
                break
        if not note_column:
            note_column = len(excel_headers) + 1
            sheet.cell(row=1, column=note_column).value = "Note"
            logging.info(f"Added 'Note' column to Excel file at column {note_column}")
            wb.save(filepath)

        if not rows:
            logging.info("No data rows to process in Excel file")
            wb.save(filepath)
            return "Success"

        terminate_chrome_processes()
        driver = initialize_driver(config)
        form_headers = get_form_headers(driver, config)
        # Save headers to a text file
        with open("form_headers.txt", "w", encoding="utf-8") as f:
            for header in form_headers:
                f.write(header + "\n")

        header_mapping, unmatched_headers = match_headers(excel_headers, form_headers)

        for idx, row in enumerate(rows, start=2):
            note_cell = sheet.cell(row=idx, column=note_column).value
            if note_cell == "Inserted":
                logging.info(f"Row {idx} already inserted, skipping")
                continue

            logging.info(f"Processing row {idx}: {row}")
            success = fill_google_form(driver, row, excel_headers, header_mapping, config)

            if success:
                sheet.cell(row=idx, column=note_column).value = "Inserted"
                logging.info(f"Row {idx} processed successfully")
            else:
                error_message = f"Failed to insert row {idx-1}: Form submission error, check field mappings or network connection"
                sheet.cell(row=idx, column=note_column).value = error_message
                logging.error(f"{error_message} - Row data: {row}")
                wb.save(filepath)
                logging.info(f"Excel file saved with error note for row {idx}")
                return error_message

            wb.save(filepath)
            logging.info(f"Excel file saved after processing row {idx}")

        wb.save(filepath)
        logging.info("Final Excel file save completed")
        return "Success"

    except Exception as e:
        error_message = f"Main process error: {e}"
        logging.error(error_message)
        if wb is not None:
            try:
                sheet = wb.active
                note_column = None
                for col_idx, cell in enumerate(sheet[1], start=1):
                    if cell.value and isinstance(cell.value, str) and cell.value.lower() == "note":
                        note_column = col_idx
                        break
                if not note_column:
                    note_column = sheet.max_column + 1
                    sheet.cell(row=1, column=note_column).value = "Note"
                sheet.cell(row=2, column=note_column).value = f"Error: {str(e)}"
                wb.save(filepath)
                logging.info("Excel file saved with error note due to critical error")
            except Exception as save_err:
                logging.error(f"Failed to save Excel file: {save_err}")
        return error_message

    finally:
        if driver:
            try:
                driver.quit()
                logging.info("WebDriver closed")
            except Exception as e:
                logging.error(f"Error closing WebDriver: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ConfigGUI(root)
    root.mainloop()
